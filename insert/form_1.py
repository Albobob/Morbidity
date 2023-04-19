import datetime
import string
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
from pprint import pprint
from Morbidity.models.form.name_of_diseases import NameOfDiseases
from Morbidity.models.geo_category import GeoCategory
from Morbidity.models.rus_total import RusTotal
from Morbidity.models.district import District
from Morbidity.models.territorial_unit import TerritorialUnit
from Morbidity.models.population_group import PopulationGroup
from Morbidity.models.type_value import TypeValue
from Morbidity.models.form.form_2 import FormTwo

abc = list(string.ascii_lowercase)
# Подключаемся к БД
engine = create_engine("sqlite:///C:/Users/SimonyanAR.FCGIE/Desktop/Project/Morbidity/database.db", echo=False)
Session = sessionmaker(bind=engine)
session = Session()


def decode(work_book, sheet, coll, rows):
    """ Функция для чтения из EXEL. На вход подается:
        work_book = файл EXEL,
        sheet = страница,
        коардинаты ячейки(колонка и строчка)
        coll = колонка,
        rows = строчка
    """
    sheet = work_book[f'{sheet}']
    value_cell = sheet[f'{coll}{rows}'].value
    return value_cell


wb = load_workbook('ЮФО.xlsx', data_only=True)
sheet = wb.sheetnames


# всё население
total = session.query(PopulationGroup).filter_by(name='total').first()
total_id = total.id
# детское население 0-17 лет (включительно)
group_5 = session.query(PopulationGroup).filter_by(name='00-17').first()
group_5 = group_5.id
# детское население 0-17 лет (включительно)
group_6 = session.query(PopulationGroup).filter_by(name='00-14').first()
group_6 = group_6.id
# детское население: новорожденные
group_2 = session.query(PopulationGroup).filter_by(name='00-01').first()
group_2 = group_2.id
# детское население: от 1 года до 2 лет
group_3 = session.query(PopulationGroup).filter_by(name='01-02').first()
group_3 = group_3.id
# детское население от 3 до 6 лет
group_4 = session.query(PopulationGroup).filter_by(name='03-06').first()
group_4 = group_4.id
type_value_abs = session.query(TypeValue).filter_by(name='Абсолютное значение').first()
type_value_abs_id = type_value_abs.id
type_value_rate = session.query(TypeValue).filter_by(name='Показатель на 100 тыс. населения').first()
type_value_rate_id = type_value_rate.id

for i in sheet:
    nod = session.query(NameOfDiseases).filter_by(krista_id=i).first()
    nod_id = nod.id
    for r in range(13):
        date = datetime.datetime(2010 + r, 1, 1)
        row = r + 11
        rg_name_xl = decode(wb, i, 'A', row)
        rg = session.query(TerritorialUnit).filter_by(name_ru='Южный федеральный округ').first()

        if rg is None:
            continue
        rg_id = rg.id
        print(rg_id)
        if i != '1.02.01.03.00':
            pass
            value = decode(wb, i, 'B', row)
            print(f'{date} {rg_id} {nod_id} {total_id} {type_value_abs_id} {value} ---- {i}')
            session.add(FormTwo(date, rg_id, nod_id, total_id, type_value_abs_id, value))
            value = decode(wb, i, 'C', row)
            session.add(FormTwo(date, rg_id, nod_id, total_id, type_value_rate_id, value))
            #
            # value = decode(wb, i, 'D', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_5, type_value_abs_id, value))
            # value = decode(wb, i, 'E', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_5, type_value_rate_id, value))
            #
            # value = decode(wb, i, 'F', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_6, type_value_abs_id, value))
            # value = decode(wb, i, 'G', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_6, type_value_rate_id, value))
            #
            # value = decode(wb, i, 'H', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_2, type_value_abs_id, value))
            # value = decode(wb, i, 'I', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_2, type_value_rate_id, value))

            # value = decode(wb, i, 'J', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_3, type_value_abs_id, value))
            # value = decode(wb, i, 'K', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_3, type_value_rate_id, value))
            #
            # value = decode(wb, i, 'L', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_4, type_value_abs_id, value))
            # value = decode(wb, i, 'M', row)
            # session.add(FormTwo(date, rg_id, nod_id, group_4, type_value_rate_id, value))

    # rec(i, row, )
    # print(f'{row - 13}| {rg_name_xl} - {rg_id}')

session.commit()
session.close()
