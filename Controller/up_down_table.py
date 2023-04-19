from datetime import datetime
import string
from sqlalchemy import create_engine, func
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
from pprint import pprint
from Morbidity.models.form.name_of_diseases import NameOfDiseases
from Morbidity.models.geo_category import GeoCategory
from Morbidity.models.rus_total import RusTotal
from Morbidity.models.district import District
from Morbidity.models.territorial_unit import TerritorialUnit
from Morbidity.models.population_group import PopulationGroup
from Morbidity.models.type_value import TypeValue
from Morbidity.models.form.form_2 import FormTwo
from Morbidity.Controller.SMP import get_smp
from openpyxl import load_workbook

engine = create_engine("sqlite:///C:/Users/SimonyanAR.FCGIE/Desktop/Project/Morbidity/database.db", echo=False)
Session = sessionmaker(bind=engine)
session = Session()

##############
# TODO: При выборе региона он автамотически вытаскивает данные по Округу и РФ
df = 2010
dl = 2023
nz_input = 29
rg_1 = 90
rg_2 = 91
rg_3 = 3


########

def get_data(date_first: int, date_last: int, nz_input: int, rg_id: int) -> list:
    df = datetime.strptime(f"{date_first}-01-01", "%Y-%m-%d")
    dl = datetime.strptime(f"{date_last}-01-01", "%Y-%m-%d")
    value = session.query(FormTwo.value).filter(
        FormTwo.rg_id == rg_id,
        FormTwo.date.between(df, dl),
        FormTwo.nod_id == nz_input,
        FormTwo.pg_id == 1,
        FormTwo.type_value == 2
    ).all()
    return [i[0] for i in value if i[0] is not None]


def record(work_book, sheet, coll, rows, value):
    """ Функция для записи в EXEL. На вход подается:
        work_book = файл EXEL,
        sheet = страница,
        коардинаты ячейки(колонка и строчка)
        coll = колонка,
        rows = строчка
        значение = value (value преобразуется в тип float)
    """
    sheet = work_book[f'{sheet}']
    sheet[f'{coll}{rows}'] = value


def get_up_down(date_first: int, date_last: int, nz_id: int, rg_id_1: int, rg_id_2: int, rg_id_3: int):
    output_name = f'Up_down_{date_first}_{date_last}_{rg_id_1}_{rg_id_2}_{rg_id_3}__{nz_id}.xlsx'
    data_1 = get_data(date_first, date_last, nz_id, rg_id_1)
    data_2 = get_data(date_first, date_last, nz_id, rg_id_2)
    data_3 = get_data(date_first, date_last, nz_id, rg_id_3)

    wb = load_workbook(filename='C:/Users/SimonyanAR.FCGIE/Desktop/Project/Morbidity/Controller/tmp.xlsx')
    abc = list(string.ascii_lowercase)

    years = [year for year in range(date_first, date_last + 1)]
    columns = [c for c in abc[2: len(years) + 1]]

    for i in range(len(columns)):
        record(wb, 'tmp', columns[i], 4, years[i])
        record(wb, 'tmp', columns[i], 5, data_1[i])
        record(wb, 'tmp', columns[i], 6, data_2[i])
        record(wb, 'tmp', columns[i], 7, data_3[i])

    tu_name_1 = session.query(TerritorialUnit.name_ru).filter(TerritorialUnit.id == rg_id_1).first()
    record(wb, 'tmp', 'b', 5, tu_name_1[0])
    tu_name_2 = session.query(TerritorialUnit.name_ru).filter(TerritorialUnit.id == rg_id_2).first()
    record(wb, 'tmp', 'b', 6, tu_name_2[0])
    tu_name_3 = session.query(TerritorialUnit.name_ru).filter(TerritorialUnit.id == rg_id_3).first()
    record(wb, 'tmp', 'b', 7, tu_name_3[0])

    smp_1 = round(get_smp(2010, 2019, nz_id, rg_id_1), 3)
    record(wb, 'tmp', 'q', 5, smp_1)
    smp_2 = round(get_smp(2010, 2019, nz_id, rg_id_2), 3)
    record(wb, 'tmp', 'q', 6, smp_2)
    smp_3 = round(get_smp(2010, 2019, nz_id, rg_id_3), 3)
    record(wb, 'tmp', 'q', 7, smp_3)

    nz_name = session.query(NameOfDiseases.name).filter(NameOfDiseases.id == nz_id).first()
    record(wb, 'tmp', 'B', 3, nz_name[0])

    wb.save(output_name)
    return output_name


