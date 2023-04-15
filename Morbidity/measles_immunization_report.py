import datetime
import string
import config
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from openpyxl import load_workbook
from models.population_group import PopulationGroup
from models.territorial_unit import TerritorialUnit
from models.district import District
from models.rus_total import RusTotal
from models.geo_category import GeoCategory

from models.report.measles_immunization_report.vaccinated_measles import VaccinatedMeasles

date = datetime.datetime.now()
# получаем номер недели в году
week_number = date.strftime('%U')

# Подключаемся к БД
engine = create_engine(f"{config.SQL_PATH}", echo=False)
Session = sessionmaker(bind=engine)
session = Session()

# Всего
total = session.query(PopulationGroup).filter_by(name='total').first()
total_id = total.id
# Дети
children = session.query(PopulationGroup).filter_by(name='00-17').first()
children_id = children.id
# Взрослые
adult = session.query(PopulationGroup).filter_by(name='18+').first()
adult_id = adult.id
# Мигранты
migrants = session.query(PopulationGroup).filter_by(name='мигранты').first()
migrants_id = migrants.id
# Пребывающие из новых территорий РФ
new_citizens = session.query(PopulationGroup).filter_by(name='переселенцы').first()
new_citizens_id = new_citizens.id

# Открываем XL
wb = load_workbook(filename='52.xlsx', data_only=True)
abc = list(string.ascii_lowercase)


def decode(work_book, sheet, coll, rows):
    """ Функция для чтения из EXEL. На вход подается:
        work_book = файл EXEL,
        sheet = страница,
        коардинаты ячейки(колонка и строчка)
        coll = колонка,
        rows = строчка
    """
    sheet = work_book[f'{sheet}']
    value_cell = sheet[f'{coll}{rows}'].value
    return value_cell


# TODO Добавить описание
def true_region_name(name: str) -> str:
    true_name = None
    if name == 'Москва':
        true_name = 'г. Москва'
        return true_name

    if name == 'Санкт-Петербург':
        true_name = 'г.Санкт-Петербург'
        return true_name

    if name == 'Республика Адыгея (Адыгея)':
        true_name = 'Республика Адыгея'
        return true_name

    if name == 'Кемеровская область - Кузбасс':
        true_name = 'Кемеровская область'
        return true_name

    if name == 'Ханты-Мансийский автономный округ-Югра' or 'Ханты-Мансийский АО':
        true_name = 'Ханты-Мансийский автономный округ — Югра'
        return true_name

    if name == 'Ямало-Ненецкий АО':
        true_name = 'Ямало-Ненецкий автономный округ'
        return true_name

    if name == 'Чувашская Республика – Чувашия':
        true_name = 'Чувашская Республика'
        return true_name

    if name == 'Республика Татарстан (Татарстан)':
        true_name = 'Республика Татарстан'
        return true_name

    if name == 'Республика Северная Осетия-Алания' or 'Республика Северная Осетия':
        true_name = 'Республика Северная Осетия — Алания'
        return true_name

    if name == 'Республика Саха ( Якутия )':
        true_name = 'Республика Саха (Якутия)'
        return true_name

    if name == 'Еврейская АО':
        true_name = 'Еврейская автономная область'
        return true_name

    if name == 'Чукотский АО':
        true_name = 'Чукотский автономный округ'
        return true_name


def insert(rg_id, rows, date_now):
    sheet = date_now.strftime('%U')

    total_unvac = decode(wb, f'{sheet}', abc[2], rows)
    total_vac_subjects = decode(wb, f'{sheet}', abc[5], rows)
    total_vac_measles = decode(wb, f'{sheet}', abc[10], rows)
    session.add(VaccinatedMeasles(rg_id, date_now, total_id, total_vac_measles, total_unvac, total_vac_subjects))

    children_unvac = decode(wb, f'{sheet}', abc[3], rows)
    children_vac_subjects = decode(wb, f'{sheet}', abc[6], rows)
    children_vac_measles = decode(wb, f'{sheet}', abc[11], rows)
    session.add(
        VaccinatedMeasles(rg_id, date_now, children_id, children_vac_measles, children_unvac, children_vac_subjects))

    adult_unvac = decode(wb, f'{sheet}', abc[4], rows)
    adult_vac_subjects = decode(wb, f'{sheet}', abc[7], rows)
    adult_vac_measles = decode(wb, f'{sheet}', abc[12], rows)
    session.add(VaccinatedMeasles(rg_id, date_now, adult_id, adult_vac_measles, adult_unvac, adult_vac_subjects))

    migrants_unvac = None
    migrants_vac_subjects = decode(wb, f'{sheet}', abc[8], rows)
    migrants_vac_measles = decode(wb, f'{sheet}', abc[13], rows)
    session.add(
        VaccinatedMeasles(rg_id, date_now, migrants_id, migrants_vac_measles, migrants_unvac, migrants_vac_subjects))

    new_citizens_unvac = None
    new_citizens_vac_subjects = decode(wb, f'{sheet}', abc[9], rows)
    new_citizens_vac_measles = decode(wb, f'{sheet}', abc[14], rows)
    session.add(
        VaccinatedMeasles(rg_id, date_now, new_citizens_id, new_citizens_vac_measles, new_citizens_unvac,
                          new_citizens_vac_subjects))


worksheet = wb[f'{week_number}']
max_row = worksheet.max_row

for row in range(8, max_row - 8):

    name = decode(wb, f'{week_number}', 'B', row)
    name_bd = session.query(TerritorialUnit).filter_by(name_ru=f'{name}').first()
    if name_bd is None:
        region_name = true_region_name(name)
    else:
        region_name = name
    try:
        region = session.query(TerritorialUnit).filter_by(name_ru=f'{region_name}').first()
        rg_id = region.id
        insert(rg_id, row, date)
    except AttributeError:
        print(f"EROR - {region_name}")

###########################
session.commit()
session.close()
